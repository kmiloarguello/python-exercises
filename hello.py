# Tkinter lib to create user interface
import sys
from Tkinter import *
from tkFileDialog import askopenfilename
from tkintertable import TableCanvas, TableModel
import tkMessageBox


# Openpyxl libs
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell


# Functions

def myFunction():
  ne = Label(window, text='Hola',bg="red")
  ne.pack()

def editMenu():
  print 'Again'

def helpMenu():
  print 'Help'

def openExcel():
  print 'hola'

def acercaDe():
  myWindow = Toplevel(window)
  myWindow.title('K@PTA')
  myWindow.wm_iconbitmap('img/kapta_mex.ico')
  myWindow.geometry('200x100')
  acercaDeContent = Label(myWindow, text='Derechos Reservados K@PTA')
  acercaDeContent.pack()

def exitApp():
  exited = tkMessageBox.askyesno('Salir','Esta seguro?')
  if(exited == True):
    window.destroy()

# Initialization

window = Tk()
window.title('K@PTA Excel Auditorias')
window.wm_iconbitmap('img/kapta_mex.ico')
window.geometry('800x600')
window.configure(background='white')

# Menu

menu = Menu(window)
window.config(menu=menu)

subMenu = Menu(menu,tearoff=0,bg='white')
menu.add_cascade(label='Archivo', menu=subMenu)
subMenu.add_command(label='Nuevo proyecto', command=myFunction)
subMenu.add_command(label='Abrir Excel', command=openExcel)
subMenu.add_command(label='Guardar', command=myFunction)
subMenu.add_command(label='Exportar Excel', command=myFunction)
subMenu.add_separator()
subMenu.add_command(label='Acerca de K@PTA', command=acercaDe)
subMenu.add_command(label='Salir', command=exitApp)

editMenu = Menu(menu,tearoff=0,bg='white')
menu.add_cascade(label='Editar', menu=editMenu)
editMenu.add_command(label='Deshacer', command=myFunction)

helpMenu = Menu(menu,tearoff=0,bg='white')
menu.add_cascade(label='Ayuda', menu=helpMenu)

# Bottom
toolbar = Frame(window,bg='white')
myLabel = Label(toolbar, text='Derechos Reservados K@PTA', bg='white')
myLabel.pack(side=RIGHT)
toolbar.pack(side=BOTTOM, fill=X)
  
window.filename = askopenfilename( filetypes = (("Archivos de Auditorias", ".xlsx"), ("Todos los archivos", "*.*")))

numberCategory = []
zerovalue = []
index_number_categories = []
rowData = []
cleanedRowData = []
auditvalue = []
essential = []

wb = load_workbook(filename = window.filename, data_only=True)
sheets = wb.sheetnames[3:12]

myHoja = []
myN = []
myZero = []
myAudit = []
myEssentials = []

for sheet in sheets:
  ws = wb[sheet]

  for row in ws.rows: 
    numberCategory.insert(0,row[23].value)  
    number_categories_without_filter = next(i for i in numberCategory if i is not None)
    index_number_categories.extend([number_categories_without_filter])

    zerovalue.insert(0,row[25].value)  
    zero_categories_without_filter = next(i for i in zerovalue if i is not None)

    auditvalue.insert(0,row[13].value)  
    audit_categories_without_filter = next(i for i in auditvalue if i is not None)

    essential.insert(0,row[15].value)  
    essential_without_filter = next(i for i in essential if i is not None)

    if(row[23].value == "N" and zero_categories_without_filter == 0):
      myHoja.extend([sheet])
      myN.extend([str(row[23].value)])
      myZero.extend([str(zero_categories_without_filter)])
      myAudit.extend([str(audit_categories_without_filter)])
      myEssentials.extend([str(essential_without_filter)])

tframe = Frame(window)
tframe.pack()
model = TableModel()
table = TableCanvas(tframe,model=model,editable=False,rowheaderwidth=50)
table.createTableFrame()
model = table.model

dict = {}

for i in range(len(myHoja)):
  dict[i] = {'ID': i}

model.importDict(dict)

table.addColumn('Hoja Excel')
table.addColumn('Standard')
table.addColumn('Number')
table.addColumn('Requirement 2015')
table.addColumn('Comments')
table.addColumn('Type of Check')
table.addColumn('Essentials')
table.addColumn('Audit Question')
table.addColumn('Observation / Evidence Required / Audit Remarks')
table.addColumn('Suggested Person to ask')
table.addColumn('Evaluation(0/1')
table.addColumn('Result')
table.addColumn('Audit Comments')
table.addColumn('Picture / Statement / Proof')

for i in range(len(myHoja)):
  table.model.data[i]['Hoja Excel'] = myHoja[i]
  # table.model.data[i]['Standard'] = 0
  # table.model.data[i]['Number'] 
  # table.model.data[i]['Requirement 2015']
  # table.model.data[i]['Comments']
  table.model.data[i]['Type of Check'] = myAudit[i]
  table.model.data[i]['Essentials'] = myEssentials[i]
  # table.model.data[i]['Audit Question']
  # table.model.data[i]['Observation / Evidence Required / Audit Remarks']
  # table.model.data[i]['Suggested Person to ask']
  table.model.data[i]['Evaluation (0/1)'] = myN[i]
  table.model.data[i]['Result'] = myZero[i]
  # table.model.data[i]['Audit Comments']
  # table.model.data[i]['Picture / Statement / Proof']


table.redrawTable()

window.mainloop()
