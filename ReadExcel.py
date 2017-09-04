from openpyxl import load_workbook
wb = load_workbook(filename = 'xlsx/BMW_Sales_Standards_2016_ME.xlsx', data_only=True)

ws = wb['Section 1_Brand Architecture']

cell_range = ws['Z6':'Z18']

for i, rowOfCellObjects in enumerate(cell_range):
  for n, cellObj in enumerate(rowOfCellObjects):
    print cellObj.value
